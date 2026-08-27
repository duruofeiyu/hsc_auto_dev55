from openpyxl import load_workbook
from openpyxl.styles import Border, Side

wb = load_workbook("reports/平台设置页测试用例.xlsx")
ws = wb.active

# 定义实线框
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# 给所有有内容的单元格加边框
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        if cell.value is not None or cell.row == 1:  # 表头即使没值也加框
            cell.border = thin_border

wb.save("reports/平台设置页测试用例.xlsx")
print("已给所有单元格添加实线边框")
