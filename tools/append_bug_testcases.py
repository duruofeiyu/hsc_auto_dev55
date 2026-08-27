from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

wb = load_workbook("reports/平台设置页测试用例.xlsx")
ws = wb.active

# 颜色定义
P0_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
P1_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
P2_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
P3_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

# 新增用例（服务器同步 bug 专项）
new_cases = [
    ["TC-HSC55-SM-031", "服务器同步-添加IP地址提示与列表状态一致性", "P1", "系统管理-平台设置", "功能测试",
     "1. 已使用管理员账号登录HSC系统\n2. 已进入【系统管理-平台设置】页面\n3. 已打开【修改系统时间】弹窗并切换到【服务器同步】tab",
     "服务器地址：192.168.1.100",
     "1. 点击【新增】按钮\n2. 输入服务器地址：192.168.1.100\n3. 保存\n4. 观察前端提示\n5. 查看服务器地址下拉列表",
     '1. 若后端返回成功：提示"添加成功"，列表显示该IP\n2. 若后端返回失败：提示"添加失败"，列表不显示该IP\n3. 提示与实际列表状态保持一致',
     "", "", '测试发现bug：提示"添加失败"但列表仍显示该IP'],
    ["TC-HSC55-SM-032", "服务器同步-删除IP地址提示与列表状态一致性", "P1", "系统管理-平台设置", "功能测试",
     "1. 已使用管理员账号登录HSC系统\n2. 已进入【系统管理-平台设置】页面\n3. 已打开【修改系统时间】弹窗并切换到【服务器同步】tab\n4. 列表中已存在服务器地址",
     "待删除服务器地址：192.168.1.100",
     "1. 选择列表中的服务器地址\n2. 点击删除按钮\n3. 观察前端提示\n4. 查看服务器地址下拉列表\n5. 刷新页面",
     '1. 若后端返回成功：提示"删除成功"，列表移除该IP，刷新后仍不存在\n2. 若后端返回失败：提示"删除失败"，列表保留该IP，刷新后仍存在\n3. 提示与实际列表状态保持一致',
     "", "", '测试发现bug：提示"删除失败"但列表已移除，刷新后确实不存在'],
    ["TC-HSC55-SM-033", "服务器同步-完整流程验证（新增-更新-删除-刷新）", "P1", "系统管理-平台设置", "功能测试",
     "1. 已使用管理员账号登录HSC系统\n2. 已进入【系统管理-平台设置】页面",
     "服务器地址：192.168.1.100",
     "1. 打开【修改系统时间】弹窗，切换到【服务器同步】tab\n2. 新增服务器地址 192.168.1.100\n3. 开启自动同步，点击【更新】\n4. 删除该服务器地址\n5. 刷新页面\n6. 重新进入服务器同步tab",
     "1. 新增、更新、删除操作提示与实际状态一致\n2. 刷新页面后数据与操作结果一致\n3. 无状态不一致或数据残留",
     "", "", "完整流程复现前端提示与实际状态不一致问题"],
]

# 从最后一行后追加
start_row = ws.max_row + 1
for row_idx, case in enumerate(new_cases, start_row):
    priority = case[2]
    for col_idx, value in enumerate(case, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # 优先级颜色
        if col_idx == 3:
            if priority == "P0":
                cell.fill = P0_FILL
            elif priority == "P1":
                cell.fill = P1_FILL
            elif priority == "P2":
                cell.fill = P2_FILL
            elif priority == "P3":
                cell.fill = P3_FILL

    ws.row_dimensions[row_idx].height = 80

# 保存
wb.save("reports/平台设置页测试用例.xlsx")
print(f"已追加 {len(new_cases)} 条用例，当前共 {ws.max_row - 1} 条")
