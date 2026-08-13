"""
把已发现的 3 个缺陷补充到 HSC_TestCases_SystemManagement_55.xlsx
"""
import os
from openpyxl import load_workbook


def add_defect_cases():
    file_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "../reports/HSC_TestCases_SystemManagement_55.xlsx"
    )

    wb = load_workbook(file_path)
    ws = wb.active

    defect_cases = [
        [
            "TC-SM-BUG-001", "登录模块/角色管理", "接口签名",
            "缺少 x-sign 签名头时接口返回 500",
            "已登录，构造不带 x-sign/x-timestamp/x-version/x-tenant-id 的请求",
            "1. 使用 Postman 或脚本调用 POST /system/auth/login 或 POST /system/role\n2. Header 中只传 Authorization、X-Access-Token、Content-Type\n3. Payload 使用正确参数\n4. 发送请求",
            "如果后端需要签名校验，应返回 401/400 并明确提示缺少签名头",
            "P1",
            "实际返回：{\"success\":false,\"code\":500,\"message\":\"抱歉，服务器开小差了\"}"
        ],
        [
            "TC-SM-BUG-002", "角色管理", "菜单权限回显",
            "新增角色全选菜单权限后，编辑时「全选/全不选」未自动勾选",
            "已创建角色，且该角色已勾选全部菜单权限",
            "1. 进入角色管理\n2. 点击新增，填写角色名称、权限字符\n3. 勾选「全选/全不选」，确认所有菜单已勾选\n4. 点击确定保存\n5. 在列表中点击该角色的「编辑」\n6. 查看「全选/全不选」复选框状态",
            "编辑回显时，若该角色已勾选全部菜单，「全选/全不选」应自动勾选",
            "P2",
            "实际：所有菜单项已勾选，但顶部「全选/全不选」复选框未勾选"
        ],
        [
            "TC-SM-BUG-003", "部门管理", "列设置",
            "列设置中勾选「排序」后，列表未显示排序列",
            "已进入部门管理页面，列表中未显示「排序」列",
            "1. 进入系统管理-组织权限-部门管理\n2. 查看当前列表显示列\n3. 点击右上角列设置图标\n4. 勾选「排序」\n5. 查看列表是否显示排序列",
            "列设置中勾选「排序」后，部门列表应显示「排序」列及对应排序值",
            "P2",
            "实际：列设置中「排序」已勾选，但列表中未显示排序列"
        ],
    ]

    for case in defect_cases:
        ws.append(case)
        row = ws.max_row
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.alignment = cell.alignment.copy(wrapText=True, vertical="top")
            # 缺陷用例标红
            cell.font = cell.font.copy(color="FF0000")

    wb.save(file_path)
    print(f"✅ 已追加 {len(defect_cases)} 条缺陷用例到：{file_path}")


if __name__ == "__main__":
    add_defect_cases()
